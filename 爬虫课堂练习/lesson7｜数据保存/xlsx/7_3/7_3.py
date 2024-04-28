import openpyxl

# 写入的代码：
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'new title'
sheet['A1'] = '童话大王'
rows = [['皮皮鲁', '鲁西西', '舒克', '贝塔'], ['是', '郑渊洁', '童话', '经典', '人物']]
for i in rows:
    sheet.append(i)
print(rows)
wb.save('Zheng.xlsx')

# 读取的代码：
# 调用load_workbook打开Zheng.xlsx
wb = openpyxl.load_workbook('Zheng.xlsx')
# 获取名为new title的工作表
sheet = wb['new title']
# 获取工作簿所有工作表的名字
sheetname = wb.sheetnames
# 打印工作表名字
print(sheetname)
# 把“new title”工作表中A1单元格赋值给A1_cell
A1_cell = sheet['A1']
# 获取A1单元格的值
A1_value = A1_cell.value
# 打印A1单元格的值
print(A1_value)
